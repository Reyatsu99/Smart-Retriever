from __future__ import annotations

from pathlib import Path


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix == ".docx":
        return _extract_docx(path)
    if suffix == ".xlsx":
        return _extract_xlsx(path)
    return ""


def _extract_pdf(path: Path) -> str:
    try:
        import pymupdf  # type: ignore # PyMuPDF
        with pymupdf.open(str(path)) as doc:
            return "\n".join(page.get_text() for page in doc)
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore
            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception:
            return ""
    except Exception:
        return ""


def _extract_docx(path: Path) -> str:
    try:
        import docx  # type: ignore
    except ImportError:
        return ""
    document = docx.Document(str(path))
    return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text)


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return ""
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        rows.append(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) for cell in row if cell is not None]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows)
